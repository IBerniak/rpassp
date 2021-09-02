'''
Working with excel table is implemented here by the class ExcelWB
'''

from openpyxl import load_workbook

class ExcelWB:
    '''
    Read and write an excel workbook with a given path.

    It's expected to have either one or more tables in the workbook:
      1. CheckList (required)
      2...n. Any table except FoundActs
    If FoundActs-table exists, it is raising an exception.

    Fields of CheckList-table:
      A2...An --  LastName
      B2...Bn -- FirstName
      C2...Cn -- Patronymic
      D2...Dn -- BirthDate
    The first row of each field (column) is a name of that field.

    Methods defined here:
    1. read_list -- reading information from the table 'CheckList'
    2. write_acts -- creating the table 'FoundActs', writing found information
    into it.

    !Working with reading table errors is not handled! Method validate_wb is in
    the future.
    '''
    def __init__(self, path='./Source/actual.xlsx'):
        self.path = path


    def read_list(self):
        '''
        Reading the data from table CheckList. The four columns are expected:
        LastName, FirstName, Patronymic, BirthDate. Names of these fields are not
        important except the LastName field, it should take a place at A1 cell,
        other names should be at the first row, too.
        '''
        sheet = load_workbook(filename = self.path)['CheckList']
        res_list = []

        for row in sheet.values:
            if row[0]:
                if row[0] != 'LastName':
                    ln, fn, pat, bd = row[0], row[1], row[2], row[3]
                    day = str(bd.day)
                    day = day if len(day) == 2 else '0'+ day
                    month = str(bd.month)
                    month = month if len(month) == 2 else '0'+ month
                    year = str(bd.year)
                    bd = f'{day}.{month}.{year}'

                    res_list.append((ln, fn, pat, bd)) #
            else:
                break

        return res_list

    def write_acts(self, list_of_acts_tuples):
        '''
        Write collected information (given as an arg in list-of-tupples-form)
        into the <name>.xlsx file.
        The file mustn't have a table 'FoundActs' else an exception raises.
        '''
        workbook = load_workbook(filename = self.path)

        try:
            sheet = workbook['FoundActs']

        except KeyError:
            sheet = workbook.create_sheet(title='FoundActs')

            sheet['A1'] = 'Должник (физ. лицо: ФИО, дата и место рождения; юр. \
            лицо: наименование, юр. адрес, фактический адрес)'
            sheet['B1'] = 'Исполнительное производство (номер, дата возбуждения)'
            sheet['C1'] = 'Реквизиты исполнительного документа (вид, дата \
            принятия органом, номер, наименование органа, выдавшего \
            исполнительный документ)'
            sheet['D1'] = 'Дата, причина окончания или прекращения ИП (статья, \
            часть, пункт основания)'
            sheet['E1'] = 'Предмет исполнения, сумма непогашенной задолженности'
            sheet['F1'] = 'Отдел судебных приставов (наименование, адрес)'
            sheet['G1'] = 'Судебный пристав-исполнитель, телефон для получения \
            информации'

            for (data, row) in zip(list_of_acts_tuples,
                                   range(2, len(list_of_acts_tuples)+2)):
                col = 1
                for rec in data:
                    input = sheet.cell(column=col, row=row, value=rec)
                    col += 1
            workbook.save(filename = self.path)

        else:
            raise IOError('Probably the workbook is used already, the new \
            one shouldn\'t have the FoundActs-sheet')
